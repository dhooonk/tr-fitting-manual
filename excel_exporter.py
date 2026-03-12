import openpyxl
from openpyxl.styles import Font, Alignment
from data_model import LibFile

def export_lib_to_excel(lib_file: LibFile, default_path: str = "lib_export.xlsx") -> str:
    """
    LibFile 객체의 데이터를 Excel 파일(.xlsx)로 내보내는 함수입니다.
    
    생성되는 시트 구조:
    1. Matrix View (행렬 뷰): 
       - 가로축(열) = 모델명
       - 세로축(행) = 파라미터명
       - 각 LIB 블록마다 이 행렬 구조가 반복되어 표시됩니다.
    2. List View (리스트 뷰): 
       - Library / Model / Type / Parameter / Value 형태의 1차원 표 구조로 모든 데이터를 나열합니다.
       
    Args:
        lib_file (LibFile): 파싱된 Smart Spice LIB 파일 데이터가 담긴 객체
        default_path (str): 저장할 Excel 파일의 기본 경로 및 이름
        
    Returns:
        str: 실제 저장된 파일의 경로 반환
    """
    wb = openpyxl.Workbook()
    
    # ── 시트 1: 행렬 뷰 (Matrix View) ──
    ws_matrix = wb.active
    ws_matrix.title = "Matrix View"
    
    current_row = 1
    
    # 스타일용
    header_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    
    for lib_block in lib_file.lib_blocks:
        if not lib_block.models:
            continue
            
        # 1. 현 LIB 블록 헤더 쓰기 (예: "LIB: NMOS_LIB")
        cell = ws_matrix.cell(row=current_row, column=1, value=f"LIB: {lib_block.name}")
        cell.font = Font(bold=True, size=12)
        current_row += 1
        
        start_row_for_lib = current_row
        
        # 모델 목록 (열 헤더)
        models = lib_block.models
        
        # 행(Row): 1열에는 'Parameter', 2열부터는 각 모델의 이름 기입
        ws_matrix.cell(row=current_row, column=1, value="Parameter").font = header_font
        for col_idx, model in enumerate(models, start=2):
            c = ws_matrix.cell(row=current_row, column=col_idx, value=model.name)
            c.font = header_font
            c.alignment = center_align
            
        current_row += 1
        
        # 2. 이 LIB 내의 모든 파라미터 이름 수집
        # 각 모델마다 가지고 있는 파라미터가 다를 수 있으므로, 중복을 제거하면서 순서를 유지해 수집합니다.
        param_names = []
        for model in models:
            for p_name in model.params.keys():
                if p_name not in param_names:
                    param_names.append(p_name)
                    
        # 3. 행별로 파라미터 값 채워넣기
        # 각 파라미터 행에 대해 반복하면서, 모델이 해당 파라미터를 가지고 있으면 값을 기입하고 없으면 빈 칸("") 처리합니다.
        for p_name in param_names:
            ws_matrix.cell(row=current_row, column=1, value=p_name)
            for col_idx, model in enumerate(models, start=2):
                val = model.params.get(p_name, "")
                ws_matrix.cell(row=current_row, column=col_idx, value=val)
            current_row += 1
            
        current_row += 2  # 다음 LIB를 위해 빈 줄
        
        
    # ── 시트 2: 리스트 뷰 (List View) ──
    ws_list = wb.create_sheet(title="List View")
    
    # 리스트 뷰 헤더 작성
    headers = ["Library", "Model", "Type", "Parameter", "Value"]
    for col_idx, h in enumerate(headers, start=1):
        c = ws_list.cell(row=1, column=col_idx, value=h)
        c.font = header_font
        
    list_row = 2
    # 전체 LIB 블록과 각 블록 안의 모델들을 순회하여 (파라미터=값) 데이터를 1줄씩 작성
    for lib_block in lib_file.lib_blocks:
        for model in lib_block.models:
            for p_name, p_val in model.params.items():
                ws_list.cell(row=list_row, column=1, value=lib_block.name)
                ws_list.cell(row=list_row, column=2, value=model.name)
                ws_list.cell(row=list_row, column=3, value=model.model_type)
                ws_list.cell(row=list_row, column=4, value=p_name)
                ws_list.cell(row=list_row, column=5, value=p_val)
                list_row += 1
                
    # 제일 긴 문자열 길이에 맞춰 Excel 시트의 열(Column) 너비를 자동으로 조정합니다.
    for ws in [ws_matrix, ws_list]:
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[col_letter].width = adjusted_width

    wb.save(default_path)
    return default_path
